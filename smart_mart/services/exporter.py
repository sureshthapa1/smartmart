"""Exporter service — PDF and CSV report exports."""

from __future__ import annotations

import csv
import io
from io import BytesIO


def export_report_csv(report_data: list[dict], columns: list[str]) -> str:
    """Return CSV string for report_data with given column headers."""
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(report_data)
    return output.getvalue()


def export_report_pdf(report_data: list[dict], title: str, columns: list[str]) -> bytes:
    """Return PDF bytes for a tabular report using ReportLab."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story = [Paragraph(title, styles["Title"]), Spacer(1, 0.4*cm)]

    table_data = [columns]
    for row in report_data:
        table_data.append([str(row.get(col, "")) for col in columns])

    col_width = (landscape(A4)[0] - 3*cm) / max(len(columns), 1)
    table = Table(table_data, colWidths=[col_width] * len(columns))
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#343a40")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
    ]))
    story.append(table)
    doc.build(story)
    return buffer.getvalue()


def export_invoice_pdf(sale) -> bytes:
    """Return PDF invoice bytes for a Sale object."""
    from smart_mart.services.sales_manager import generate_invoice_pdf
    return generate_invoice_pdf(sale.id)


def export_report_excel(report_data: list[dict], title: str, columns: list[str]) -> bytes:
    """Return Excel (.xlsx) bytes for a tabular report using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit

    # Title row
    ws.merge_cells(f"A1:{chr(64+len(columns))}1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Header row
    header_fill = PatternFill("solid", fgColor="1e3a5f")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row in enumerate(report_data, 3):
        for col_idx, col_name in enumerate(columns, 1):
            val = row.get(col_name, "")
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-width columns
    for col in ws.columns:
        try:
            max_len = max((len(str(cell.value or "")) for cell in col if hasattr(cell, 'value')), default=10)
            if hasattr(col[0], 'column_letter'):
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        except Exception:
            pass

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
