"""Purchases blueprint — purchase creation, listing, and supplier management."""

from __future__ import annotations

from datetime import date

from flask import Blueprint, flash, redirect, render_template, request, url_for
from flask_login import current_user

from ...extensions import db
from ...models.product import Product
from ...services import purchase_manager
from ...services.decorators import login_required, admin_required

purchases_bp = Blueprint("purchases", __name__, url_prefix="/purchases")


@purchases_bp.route("/")
@login_required
def list_purchases():
    start_date_raw = request.args.get("start_date", "").strip() or None
    end_date_raw = request.args.get("end_date", "").strip() or None
    filters: dict = {}
    if start_date_raw:
        try:
            filters["start_date"] = date.fromisoformat(start_date_raw)
        except ValueError:
            flash("Invalid start date format.", "danger")
    if end_date_raw:
        try:
            filters["end_date"] = date.fromisoformat(end_date_raw)
        except ValueError:
            flash("Invalid end date format.", "danger")
    page = int(request.args.get("page", 1))
    purchases = purchase_manager.list_purchases(filters, page=page)
    return render_template("purchases/list.html", purchases=purchases,
                           start_date=start_date_raw or "", end_date=end_date_raw or "")


@purchases_bp.route("/create", methods=["GET", "POST"])
@login_required
def create_purchase():
    suppliers = purchase_manager.list_suppliers()
    products = db.session.execute(db.select(Product).order_by(Product.name)).scalars().all()
    if request.method == "POST":
        supplier_id_raw = request.form.get("supplier_id", "").strip()
        purchase_date_raw = request.form.get("purchase_date", "").strip()
        items = _parse_items(request.form)
        errors = []
        if not supplier_id_raw:
            errors.append("Please select a supplier.")
        if not purchase_date_raw:
            errors.append("Please provide a purchase date.")
        if not items:
            errors.append("Please add at least one item.")
        purchase_date = None
        if purchase_date_raw:
            try:
                purchase_date = date.fromisoformat(purchase_date_raw)
            except ValueError:
                errors.append("Invalid purchase date format.")
        if errors:
            for msg in errors:
                flash(msg, "danger")
            return render_template("purchases/create.html", suppliers=suppliers, products=products,
                                   today=date.today().isoformat())
        try:
            purchase = purchase_manager.create_purchase(
                supplier_id=int(supplier_id_raw), items=items,
                purchase_date=purchase_date, user_id=current_user.id,
            )
            flash(f"Purchase #{purchase.id} created successfully.", "success")
            return redirect(url_for("purchases.list_purchases"))
        except ValueError as e:
            flash(str(e), "danger")
        except Exception as e:
            flash(f"Error creating purchase: {e}", "danger")
    return render_template("purchases/create.html", suppliers=suppliers, products=products,
                           today=date.today().isoformat())


@purchases_bp.route("/suppliers")
@login_required
def list_suppliers():
    suppliers = purchase_manager.list_suppliers()
    return render_template("purchases/suppliers.html", suppliers=suppliers)


@purchases_bp.route("/suppliers/create", methods=["GET", "POST"])
@admin_required
def create_supplier():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        contact = request.form.get("contact", "").strip()
        email = request.form.get("email", "").strip()
        address = request.form.get("address", "").strip()
        if not name:
            flash("Supplier name is required.", "danger")
            return render_template("purchases/supplier_form.html", supplier=None)
        try:
            supplier = purchase_manager.create_supplier({
                "name": name, "contact": contact or None,
                "email": email or None, "address": address or None,
            })
            flash(f"Supplier '{supplier.name}' created successfully.", "success")
            return redirect(url_for("purchases.list_suppliers"))
        except Exception as e:
            flash(f"Error creating supplier: {e}", "danger")
    return render_template("purchases/supplier_form.html", supplier=None)


@purchases_bp.route("/suppliers/<int:supplier_id>/edit", methods=["GET", "POST"])
@admin_required
def edit_supplier(supplier_id):
    from ...models.supplier import Supplier
    from ...extensions import db as _db
    supplier = _db.get_or_404(Supplier, supplier_id)
    if request.method == "POST":
        data = {
            "name": request.form.get("name", "").strip(),
            "contact": request.form.get("contact", "").strip() or None,
            "email": request.form.get("email", "").strip() or None,
            "address": request.form.get("address", "").strip() or None,
        }
        try:
            purchase_manager.update_supplier(supplier_id, data)
            flash("Supplier updated successfully.", "success")
            return redirect(url_for("purchases.list_suppliers"))
        except Exception as e:
            flash(str(e), "danger")
    return render_template("purchases/supplier_form.html", supplier=supplier)


@purchases_bp.route("/suppliers/<int:supplier_id>/delete", methods=["POST"])
@admin_required
def delete_supplier(supplier_id):
    try:
        purchase_manager.delete_supplier(supplier_id)
        flash("Supplier deleted.", "success")
    except ValueError as e:
        flash(str(e), "danger")
    return redirect(url_for("purchases.list_suppliers"))


@purchases_bp.route("/download-sample")
@login_required
def download_sample():
    """Download a sample CSV template for bulk upload."""
    from flask import Response
    sample = "Product Name,Quantity,Unit Cost,SKU,Category\n"
    sample += "Basmati Rice,50,120.00,RICE-001,Grains & Pulses\n"
    sample += "Mustard Oil (1L),30,180.00,OIL-001,Oils & Fats\n"
    sample += "Colgate Toothpaste,20,95.00,TOOTH-001,Personal Care & Hygiene\n"
    sample += "Wai Wai Noodles,100,25.00,NOODLE-001,Snacks & Bakery\n"
    sample += "Mineral Water (1L),200,20.00,WATER-001,Beverages\n"
    return Response(
        sample,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=bulk_purchase_sample.csv"}
    )


@purchases_bp.route("/bulk-upload", methods=["GET", "POST"])
@login_required
def bulk_upload():
    """Bulk purchase upload via CSV or Excel file."""
    import io, csv
    suppliers = purchase_manager.list_suppliers()

    if request.method == "POST":
        supplier_id_raw = request.form.get("supplier_id", "").strip()
        purchase_date_raw = request.form.get("purchase_date", "").strip()
        file = request.files.get("file")

        errors = []
        if not supplier_id_raw:
            errors.append("Please select a supplier.")
        if not purchase_date_raw:
            errors.append("Please provide a purchase date.")
        if not file or file.filename == "":
            errors.append("Please upload a CSV or Excel file.")

        purchase_date = None
        if purchase_date_raw:
            try:
                purchase_date = date.fromisoformat(purchase_date_raw)
            except ValueError:
                errors.append("Invalid date format.")

        if errors:
            for msg in errors:
                flash(msg, "danger")
            return render_template("purchases/bulk_upload.html", suppliers=suppliers,
                                   today=date.today().isoformat())

        # Parse file
        filename = file.filename.lower()
        rows = []
        parse_errors = []

        try:
            if filename.endswith(".csv"):
                content = file.read().decode("utf-8-sig")
                reader = csv.DictReader(io.StringIO(content))
                for i, row in enumerate(reader, 2):
                    rows.append((i, row))
            elif filename.endswith((".xlsx", ".xls")):
                import openpyxl
                wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
                ws = wb.active
                headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                    rows.append((i, dict(zip(headers, row))))
            else:
                flash("Only .csv or .xlsx files are supported.", "danger")
                return render_template("purchases/bulk_upload.html", suppliers=suppliers,
                                       today=date.today().isoformat())
        except Exception as e:
            flash(f"Error reading file: {e}", "danger")
            return render_template("purchases/bulk_upload.html", suppliers=suppliers,
                                   today=date.today().isoformat())

        # Process rows — find or create products
        items = []
        for row_num, row in rows:
            # Normalize keys
            r = {k.strip().lower(): str(v).strip() if v is not None else "" for k, v in row.items()}
            product_name = r.get("product name") or r.get("product") or r.get("name") or ""
            qty_raw = r.get("quantity") or r.get("qty") or "0"
            cost_raw = r.get("unit cost") or r.get("cost") or r.get("unit_cost") or r.get("price") or "0"
            sku_raw = r.get("sku") or r.get("barcode") or ""
            category_raw = r.get("category") or ""

            if not product_name:
                parse_errors.append(f"Row {row_num}: missing product name, skipped.")
                continue
            try:
                qty = int(float(qty_raw))
                cost = float(cost_raw)
            except ValueError:
                parse_errors.append(f"Row {row_num}: invalid quantity or cost for '{product_name}', skipped.")
                continue
            if qty <= 0:
                parse_errors.append(f"Row {row_num}: quantity must be > 0 for '{product_name}', skipped.")
                continue

            # Find existing product by name or SKU, or create new one
            product = None
            if sku_raw:
                product = db.session.execute(
                    db.select(Product).filter_by(sku=sku_raw)
                ).scalar_one_or_none()
            if not product:
                product = db.session.execute(
                    db.select(Product).filter(
                        db.func.lower(Product.name) == product_name.lower()
                    )
                ).scalar_one_or_none()
            if not product:
                # Auto-create the product with cost as both cost and selling price
                import uuid
                auto_sku = sku_raw or f"AUTO-{product_name[:6].upper().replace(' ','-')}-{str(uuid.uuid4())[:4].upper()}"
                try:
                    from ...services import inventory_manager
                    product = inventory_manager.create_product({
                        "name": product_name,
                        "category": category_raw or None,
                        "sku": auto_sku,
                        "cost_price": cost,
                        "selling_price": cost,
                        "quantity": 0,
                    })
                except Exception as e:
                    parse_errors.append(f"Row {row_num}: could not create product '{product_name}': {e}")
                    continue

            items.append({"product_id": product.id, "quantity": qty, "unit_cost": cost})

        if not items:
            flash("No valid rows found in the file. Check the format.", "danger")
            if parse_errors:
                for e in parse_errors[:5]:
                    flash(e, "warning")
            return render_template("purchases/bulk_upload.html", suppliers=suppliers,
                                   today=date.today().isoformat())

        try:
            purchase = purchase_manager.create_purchase(
                supplier_id=int(supplier_id_raw),
                items=items,
                purchase_date=purchase_date,
                user_id=current_user.id,
            )
            flash(f"✅ Bulk purchase #{purchase.id} created — {len(items)} products, "
                  f"NPR {float(purchase.total_cost):,.2f} total.", "success")
            if parse_errors:
                for e in parse_errors[:5]:
                    flash(e, "warning")
            return redirect(url_for("purchases.list_purchases"))
        except Exception as e:
            flash(f"Error saving purchase: {e}", "danger")

    return render_template("purchases/bulk_upload.html", suppliers=suppliers,
                           today=date.today().isoformat())


def _parse_items(form) -> list[dict]:
    items: list[dict] = []
    index = 0
    while True:
        product_id = form.get(f"items[{index}][product_id]")
        if product_id is None:
            break
        try:
            pid = int(product_id)
            qty = int(form.get(f"items[{index}][quantity]", "0"))
            cost = float(form.get(f"items[{index}][unit_cost]", "0"))
            if pid > 0 and qty > 0 and cost >= 0:
                items.append({"product_id": pid, "quantity": qty, "unit_cost": cost})
        except (ValueError, TypeError):
            pass
        index += 1
    return items
