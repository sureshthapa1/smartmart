"""Inventory blueprint — product CRUD, stock adjustment, and category management."""

from decimal import Decimal
import os
import uuid

from flask import Blueprint, flash, redirect, render_template, request, url_for, current_app, abort
from flask_login import current_user

from ...extensions import db
from ...models.product import Product
from ...models.supplier import Supplier
from ...models.category import Category
from ...services import inventory_manager
from ...services.decorators import admin_required, login_required

inventory_bp = Blueprint("inventory", __name__, url_prefix="/inventory")

def _require_perm(perm: str):
    from flask import abort
    from flask_login import current_user as _cu
    if _cu.role != 'admin':
        from ...models.user_permissions import UserPermissions
        p = UserPermissions.get_or_create(_cu.id)
        if not getattr(p, perm, False):
            abort(403)



ALLOWED_EXTENSIONS = {"jpg", "jpeg", "png", "gif", "webp", "bmp"}


def _allowed_file(filename: str) -> bool:
    from ...services.image_service import ALLOWED_EXTENSIONS
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def _save_product_image(file, product_name: str = None) -> str | None:
    """Upload product image via image_service (Cloudinary or local fallback)."""
    from ...services.image_service import save_product_image
    return save_product_image(file, product_name=product_name)


def _delete_product_image(identifier: str) -> None:
    """Delete product image via image_service."""
    from ...services.image_service import delete_product_image
    delete_product_image(identifier)


def _relink_product_images() -> int:
    """
    Scan uploads/products folder and auto-link images to products that
    have no image set. Matches by slugified product name in the filename.
    Returns count of products relinked.
    """
    import re as _re
    upload_dir = os.path.join(current_app.static_folder, "uploads", "products")
    if not os.path.exists(upload_dir):
        return 0

    # Build list of valid image files (non-zero)
    valid_files = []
    for f in os.listdir(upload_dir):
        if f == ".gitkeep":
            continue
        path = os.path.join(upload_dir, f)
        if os.path.getsize(path) > 100:
            valid_files.append(f)

    if not valid_files:
        return 0

    def _slugify(text: str) -> str:
        return _re.sub(r"[^a-z0-9]", "", text.lower())

    relinked = 0
    products = db.session.execute(
        db.select(Product).where(
            db.or_(Product.image_filename.is_(None), Product.image_filename == "")
        )
    ).scalars().all()

    for product in products:
        name_slug = _slugify(product.name)
        if not name_slug:
            continue
        best = None
        for fname in valid_files:
            fname_slug = _slugify(fname.rsplit(".", 1)[0])
            # Match if product name appears in filename
            if name_slug in fname_slug or fname_slug in name_slug:
                best = fname
                break
        if best:
            product.image_filename = best
            relinked += 1

    if relinked:
        db.session.commit()
    return relinked


@inventory_bp.route("/")
@login_required
def list_products():
    search = request.args.get("q", "").strip() or None
    page = request.args.get("page", 1, type=int)
    status_filter = request.args.get("status", "active")  # active | inactive | all | low

    # Map status param to active_only flag
    active_only = None
    if status_filter == "active":
        active_only = True
    elif status_filter == "inactive":
        active_only = False
    # status == "all" → active_only stays None

    if status_filter == "low":
        from sqlalchemy import func as _func, or_
        from ...models.shop_settings import ShopSettings as _SS
        try:
            _threshold = _SS.get().low_stock_threshold or 10
        except Exception:
            _threshold = 10
        term = search.strip().lower() if search else None
        stmt_products = db.select(Product).where(
            Product.is_active == True,
            Product.quantity <= db.func.coalesce(Product.low_stock_threshold, _threshold),
        ).order_by(Product.quantity.asc(), Product.name.asc())
        if term:
            stmt_products = stmt_products.where(or_(
                db.func.lower(Product.name).contains(term),
                db.func.lower(Product.category).contains(term),
                db.func.lower(Product.sku) == term,
            ))
        products = db.session.execute(
            stmt_products.limit(100).offset((page - 1) * 100)
        ).unique().scalars().all()
    else:
        products = inventory_manager.get_products(search=search, page=page, active_only=active_only)

    # Total count for pagination (same filters)
    from sqlalchemy import func as _func, or_
    stmt = db.select(_func.count(Product.id))
    if search:
        term = search.strip().lower()
        stmt = stmt.where(
            or_(
                db.func.lower(Product.name).contains(term),
                db.func.lower(Product.category).contains(term),
                db.func.lower(Product.sku) == term,
            )
        )
    if active_only is True:
        stmt = stmt.where(Product.is_active == True)
    elif active_only is False:
        stmt = stmt.where(Product.is_active == False)
    if status_filter == "low":
        stmt = stmt.where(Product.is_active == True)
        stmt = stmt.where(Product.quantity <= _func.coalesce(Product.low_stock_threshold, _threshold))

    total = db.session.execute(stmt).scalar() or 0
    per_page = 100
    total_pages = max(1, (total + per_page - 1) // per_page)
    price_alert_product_ids = set()
    try:
        from ...models.supplier_price_record import SupplierPriceRecord
        latest_rows = db.session.execute(
            db.select(SupplierPriceRecord)
            .where(SupplierPriceRecord.product_id.in_([p.id for p in products] or [0]))
            .order_by(SupplierPriceRecord.product_id, SupplierPriceRecord.recorded_at.desc())
        ).scalars().all()
        grouped = {}
        for row in latest_rows:
            grouped.setdefault(row.product_id, []).append(row)
        for product_id, rows in grouped.items():
            if len(rows) >= 2 and float(rows[1].cost_price or 0) > 0:
                increase = (float(rows[0].cost_price) - float(rows[1].cost_price)) / float(rows[1].cost_price)
                if increase >= 0.10:
                    price_alert_product_ids.add(product_id)
    except Exception:
        pass

    from datetime import date as _date, datetime, timezone
    return render_template("inventory/list.html", products=products, search=search or "",
                           page=page, total=total, total_pages=total_pages, per_page=per_page,
                           status_filter=status_filter, today=_date.today(),
                           price_alert_product_ids=price_alert_product_ids)


@inventory_bp.route("/create", methods=["GET", "POST"])
@login_required
def create_product():
    if current_user.role != "admin":
        from ...models.user_permissions import UserPermissions
        p = UserPermissions.get_or_create(current_user.id)
        if not p.can_add_product:
            abort(403)
    suppliers = db.session.execute(db.select(Supplier).order_by(Supplier.name)).scalars().all()
    try:
        categories = db.session.execute(db.select(Category).order_by(Category.name)).scalars().all()
    except Exception:
        categories = []
    if request.method == "POST":
        data = _form_to_data(request.form)
        # Pass user_id and purchase_date so inventory_manager creates purchase record
        data["user_id"] = current_user.id
        # Handle image upload (admin only)
        if current_user.role == "admin":
            img_file = request.files.get("product_image")
            img_filename = _save_product_image(img_file, product_name=data.get("name"))
            if img_filename:
                data["image_filename"] = img_filename
            # Save custom emoji to ProductIconMap
            custom_emoji = request.form.get("custom_emoji", "").strip()
            if custom_emoji and data.get("name"):
                try:
                    from ...models.product_icon_map import ProductIconMap
                    ProductIconMap.set(data["name"], custom_emoji)
                except Exception:
                    pass
        try:
            inventory_manager.create_product(data)
            qty = data.get("quantity", 0)
            if qty and int(qty) > 0 and data.get("supplier_id"):
                flash(f"✅ Product added and opening stock of {qty} units recorded as a purchase.", "success")
            elif qty and int(qty) > 0:
                flash(f"✅ Product added with opening stock of {qty} units. Add a supplier to track it as a purchase expense.", "info")
            else:
                flash("✅ Product created. Add stock by recording a purchase.", "success")
            return redirect(url_for("inventory.list_products"))
        except ValueError as e:
            flash(str(e), "danger")
    return render_template("inventory/form.html", product=None, suppliers=suppliers,
                           categories=categories, action="Create",
                           today=__import__("datetime").date.today())


@inventory_bp.route("/<int:product_id>/edit", methods=["GET", "POST"])
@login_required
def edit_product(product_id):
    if current_user.role != "admin":
        from ...models.user_permissions import UserPermissions
        p = UserPermissions.get_or_create(current_user.id)
        if not p.can_edit_product:
            abort(403)
    product = db.get_or_404(Product, product_id)
    suppliers = db.session.execute(db.select(Supplier).order_by(Supplier.name)).scalars().all()
    try:
        categories = db.session.execute(db.select(Category).order_by(Category.name)).scalars().all()
    except Exception:
        categories = []
    if request.method == "POST":
        data = _form_to_data(request.form)
        # Quantity is never editable here, regardless of what the client
        # sends — stock levels must only change via a tracked Purchase
        # (restocking) or a Stock Take (correcting a count). The form's
        # readonly attribute on this field is client-side only and gives no
        # protection against a direct POST bypassing the UI.
        data["quantity"] = product.quantity
        # Handle image upload (admin only)
        if current_user.role == "admin":
            img_file = request.files.get("product_image")
            img_filename = _save_product_image(img_file, product_name=product.name)
            if img_filename:
                data["image_filename"] = img_filename
            # Handle image removal
            if request.form.get("remove_image") == "1":
                _delete_product_image(product.image_filename)
                data["image_filename"] = None
            # Save custom emoji to ProductIconMap
            custom_emoji = request.form.get("custom_emoji", "").strip()
            if custom_emoji and data.get("name"):
                try:
                    from ...models.product_icon_map import ProductIconMap
                    ProductIconMap.set(data["name"], custom_emoji)
                except Exception:
                    pass
        try:
            inventory_manager.update_product(product_id, data)
            # Auto-fill any still-empty fields
            try:
                from ...services.product_autofill import autofill_product as _autofill
                _autofill(product, force=False)
            except Exception:
                pass
            flash("Product updated successfully.", "success")
            return redirect(url_for("inventory.list_products"))
        except ValueError as e:
            flash(str(e), "danger")
    return render_template("inventory/form.html", product=product, suppliers=suppliers,
                           categories=categories, action="Edit",
                           today=__import__("datetime").date.today())


@inventory_bp.route("/<int:product_id>/delete", methods=["POST"])
@admin_required
def delete_product(product_id):
    try:
        inventory_manager.delete_product(product_id)
        flash("Product deleted successfully.", "success")
    except ValueError as e:
        flash(str(e), "danger")
    except Exception as e:
        db.session.rollback()
        flash(f"Cannot delete product: {e}", "danger")
    return redirect(url_for("inventory.list_products"))


@inventory_bp.route("/<int:product_id>/autofill", methods=["POST"])
@admin_required
def autofill_product(product_id):
    """Manually trigger auto-fill of description, image, and pack_size for a product."""
    product = db.get_or_404(Product, product_id)
    force = request.form.get("force", "0") == "1"
    try:
        from ...services.product_autofill import autofill_product as _autofill
        updated = _autofill(product, force=force)
        if updated:
            fields = ", ".join(updated.keys())
            flash(f"Auto-filled: {fields} for '{product.name}'.", "success")
        else:
            flash(f"'{product.name}' already has all fields filled. Use Force to overwrite.", "info")
    except Exception as exc:
        flash(f"Auto-fill failed: {exc}", "danger")
    return redirect(url_for("inventory.edit_product", product_id=product_id))


@inventory_bp.route("/<int:product_id>/adjust-stock", methods=["GET", "POST"])
@login_required
def adjust_stock(product_id):
    if current_user.role != "admin":
        from ...models.user_permissions import UserPermissions
        p = UserPermissions.get_or_create(current_user.id)
        if not p.can_adjust_stock:
            abort(403)
    product = db.get_or_404(Product, product_id)
    if request.method == "POST":
        direction = request.form.get("direction", "in")
        note = request.form.get("note", "").strip()
        adjustment_type = request.form.get("adjustment_type", "").strip() or None
        try:
            qty = int(request.form.get("quantity", 0))
            if qty <= 0:
                raise ValueError("Quantity must be a positive integer.")
            inventory_manager.adjust_stock(product_id, qty, direction, note, current_user.id,
                                           adjustment_type=adjustment_type)
            flash(f"Stock {'added' if direction == 'in' else 'removed'} successfully.", "success")
            return redirect(url_for("inventory.list_products"))
        except ValueError as e:
            flash(str(e), "danger")
    return render_template("inventory/adjust_stock.html", product=product)


# --- Category management (admin only) ---

@inventory_bp.route("/categories")
@login_required
def list_categories():
    _require_perm("can_manage_categories")
    from sqlalchemy import func, case
    from ...models.sale import Sale, SaleItem
    from datetime import date, timedelta

    # Base category list
    cats = db.session.execute(db.select(Category).order_by(Category.name)).scalars().all()

    # Build per-category stats in one pass
    # Product counts + stock per category
    product_stats = db.session.execute(
        db.select(
            Product.category,
            func.count(Product.id).label("product_count"),
            func.coalesce(func.sum(Product.quantity), 0).label("total_stock"),
            func.coalesce(func.sum(Product.quantity * Product.cost_price), 0).label("stock_value"),
            func.coalesce(func.sum(Product.quantity * Product.selling_price), 0).label("retail_value"),
            func.sum(case((Product.quantity == 0, 1), else_=0)).label("out_of_stock"),
            func.sum(case((Product.quantity <= 10, 1), else_=0)).label("low_stock"),
        )
        .group_by(Product.category)
    ).all()
    stats_map = {r.category or "": r for r in product_stats}

    # Revenue per category (last 30 days)
    thirty_days_ago = date.today() - timedelta(days=30)
    rev_rows = db.session.execute(
        db.select(
            Product.category,
            func.coalesce(func.sum(SaleItem.subtotal), 0).label("revenue"),
            func.coalesce(func.sum(SaleItem.quantity), 0).label("qty_sold"),
        )
        .join(SaleItem, SaleItem.product_id == Product.id)
        .join(Sale, Sale.id == SaleItem.sale_id)
        .where(Sale.sale_date >= thirty_days_ago)
        .group_by(Product.category)
    ).all()
    rev_map = {r.category or "": r for r in rev_rows}

    # Combine
    category_data = []
    for cat in cats:
        s = stats_map.get(cat.name, None)
        r = rev_map.get(cat.name, None)
        category_data.append({
            "cat": cat,
            "product_count": s.product_count if s else 0,
            "total_stock": int(s.total_stock) if s else 0,
            "stock_value": float(s.stock_value) if s else 0.0,
            "retail_value": float(s.retail_value) if s else 0.0,
            "out_of_stock": int(s.out_of_stock) if s else 0,
            "low_stock": int(s.low_stock) if s else 0,
            "revenue_30d": float(r.revenue) if r else 0.0,
            "qty_sold_30d": int(r.qty_sold) if r else 0,
        })

    # Summary totals
    total_products = sum(c["product_count"] for c in category_data)
    total_stock_value = sum(c["stock_value"] for c in category_data)
    total_revenue_30d = sum(c["revenue_30d"] for c in category_data)

    return render_template("inventory/categories.html",
                           category_data=category_data,
                           total_products=total_products,
                           total_stock_value=total_stock_value,
                           total_revenue_30d=total_revenue_30d)


@inventory_bp.route("/categories/<int:cat_id>")
@login_required
def category_detail(cat_id):
    _require_perm("can_manage_categories")
    from sqlalchemy import func, case
    from ...models.sale import Sale, SaleItem
    from ...models.stock_movement import StockMovement
    from datetime import date, timedelta

    cat = db.get_or_404(Category, cat_id)
    products = db.session.execute(
        db.select(Product).where(Product.category == cat.name).order_by(Product.name)
    ).scalars().all()

    # Per-product sales stats (last 30 days)
    thirty_days_ago = date.today() - timedelta(days=30)
    sales_stats = db.session.execute(
        db.select(
            SaleItem.product_id,
            func.coalesce(func.sum(SaleItem.subtotal), 0).label("revenue"),
            func.coalesce(func.sum(SaleItem.quantity), 0).label("qty_sold"),
            func.count(SaleItem.id.distinct()).label("txn_count"),
        )
        .join(Sale, Sale.id == SaleItem.sale_id)
        .where(
            SaleItem.product_id.in_([p.id for p in products]),
            Sale.sale_date >= thirty_days_ago,
        )
        .group_by(SaleItem.product_id)
    ).all()
    sales_map = {r.product_id: r for r in sales_stats}

    # Category-level totals
    total_stock = sum(p.quantity for p in products)
    total_stock_value = sum(float(p.cost_price) * p.quantity for p in products)
    total_retail_value = sum(float(p.selling_price) * p.quantity for p in products)
    total_revenue_30d = sum(float(sales_map[p.id].revenue) if p.id in sales_map else 0 for p in products)
    total_qty_sold_30d = sum(int(sales_map[p.id].qty_sold) if p.id in sales_map else 0 for p in products)
    out_of_stock = sum(1 for p in products if p.quantity == 0)
    low_stock = sum(1 for p in products if 0 < p.quantity <= 10)

    # Daily revenue trend (last 14 days) for this category
    trend_rows = db.session.execute(
        db.select(
            func.date(Sale.sale_date).label("day"),
            func.coalesce(func.sum(SaleItem.subtotal), 0).label("revenue"),
        )
        .join(SaleItem, SaleItem.sale_id == Sale.id)
        .join(Product, Product.id == SaleItem.product_id)
        .where(
            Product.category == cat.name,
            Sale.sale_date >= date.today() - timedelta(days=13),
        )
        .group_by(func.date(Sale.sale_date))
        .order_by(func.date(Sale.sale_date))
    ).all()
    trend_labels = [str(r.day) for r in trend_rows]
    trend_data = [float(r.revenue) for r in trend_rows]

    # Enrich products with their stats
    enriched = []
    for p in products:
        s = sales_map.get(p.id)
        enriched.append({
            "product": p,
            "revenue_30d": float(s.revenue) if s else 0.0,
            "qty_sold_30d": int(s.qty_sold) if s else 0,
            "txn_count": int(s.txn_count) if s else 0,
            "profit_30d": (float(s.revenue) - float(p.cost_price) * int(s.qty_sold)) if s else 0.0,
        })
    # Sort by revenue desc
    enriched.sort(key=lambda x: x["revenue_30d"], reverse=True)

    return render_template("inventory/category_detail.html",
                           cat=cat, enriched=enriched,
                           total_stock=total_stock,
                           total_stock_value=total_stock_value,
                           total_retail_value=total_retail_value,
                           total_revenue_30d=total_revenue_30d,
                           total_qty_sold_30d=total_qty_sold_30d,
                           out_of_stock=out_of_stock,
                           low_stock=low_stock,
                           trend_labels=trend_labels,
                           trend_data=trend_data,
                           today=date.today())


@inventory_bp.route("/categories/create", methods=["GET", "POST"])
@login_required
def create_category():
    _require_perm("can_manage_categories")
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("Category name is required.", "danger")
            return render_template("inventory/category_form.html", category=None)
        existing = db.session.execute(
            db.select(Category).filter_by(name=name)
        ).scalar_one_or_none()
        if existing:
            flash(f"Category '{name}' already exists.", "danger")
            return render_template("inventory/category_form.html", category=None)
        cat = Category(name=name)
        db.session.add(cat)
        db.session.commit()
        flash(f"Category '{name}' created.", "success")
        return redirect(url_for("inventory.list_categories"))
    return render_template("inventory/category_form.html", category=None)


@inventory_bp.route("/categories/<int:cat_id>/edit", methods=["GET", "POST"])
@login_required
def edit_category(cat_id):
    _require_perm("can_manage_categories")
    cat = db.get_or_404(Category, cat_id)
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("Category name is required.", "danger")
        else:
            cat.name = name
            db.session.commit()
            flash("Category updated.", "success")
            return redirect(url_for("inventory.list_categories"))
    return render_template("inventory/category_form.html", category=cat)


@inventory_bp.route("/categories/<int:cat_id>/delete", methods=["POST"])
@login_required
def delete_category(cat_id):
    _require_perm("can_manage_categories")
    cat = db.get_or_404(Category, cat_id)
    db.session.delete(cat)
    db.session.commit()
    flash(f"Category '{cat.name}' deleted.", "success")
    return redirect(url_for("inventory.list_categories"))


# ── Bulk Product Upload ──────────────────────────────────────────────────

@inventory_bp.route("/bulk-upload", methods=["GET", "POST"])
@admin_required
def bulk_upload():
    """Bulk product upload via CSV or Excel."""
    import io, csv, uuid as _uuid

    if request.method == "POST":
        file = request.files.get("file")
        if not file or file.filename == "":
            flash("Please select a CSV or Excel file.", "danger")
            return render_template("inventory/bulk_upload.html")

        filename = file.filename.lower()
        rows = []
        try:
            if filename.endswith(".csv"):
                content = file.read().decode("utf-8-sig")
                reader = csv.DictReader(io.StringIO(content))
                for i, row in enumerate(reader, 2):
                    rows.append((i, row))
            elif filename.endswith((".xlsx", ".xls")):
                import openpyxl
                wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
                ws = wb.active
                headers = [str(c.value).strip().lower() if c.value else ""
                           for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                    rows.append((i, dict(zip(headers, row))))
            else:
                flash("Only .csv or .xlsx files are supported.", "danger")
                return render_template("inventory/bulk_upload.html")
        except Exception as e:
            flash(f"Error reading file: {e}", "danger")
            return render_template("inventory/bulk_upload.html")

        created = updated = skipped = 0
        errors = []

        for row_num, row in rows:
            r = {k.strip().lower(): str(v).strip() if v is not None else ""
                 for k, v in row.items()}

            name = r.get("product name") or r.get("name") or r.get("product") or ""
            sku = r.get("sku") or r.get("barcode") or ""
            category = r.get("category") or ""
            cost_raw = r.get("cost price") or r.get("cost") or r.get("cost_price") or "0"
            sell_raw = r.get("selling price") or r.get("price") or r.get("selling_price") or "0"
            qty_raw = r.get("quantity") or r.get("qty") or "0"
            unit = r.get("unit") or "pcs"
            supplier_name = r.get("supplier") or ""
            expiry_raw = r.get("expiry date") or r.get("expiry") or ""

            if not name:
                errors.append(f"Row {row_num}: missing product name, skipped.")
                skipped += 1
                continue

            try:
                # Strip commas, currency symbols and whitespace so values like
                # "1,500", "NPR 850", "Rs.400" or "₹ 200" parse correctly
                def _clean_num(raw: str) -> str:
                    import re as _re
                    return _re.sub(r"[^\d.\-]", "", raw.replace(",", ""))

                cost = float(_clean_num(cost_raw)) if cost_raw else 0.0
                sell = float(_clean_num(sell_raw)) if sell_raw else cost
                qty  = int(float(_clean_num(qty_raw))) if qty_raw else 0
            except ValueError:
                errors.append(f"Row {row_num}: invalid number for '{name}', skipped.")
                skipped += 1
                continue

            # Auto-generate SKU if missing
            if not sku:
                sku = f"{name[:4].upper().replace(' ', '')}-{_uuid.uuid4().hex[:4].upper()}"

            # Resolve supplier — create if not found
            supplier_id = None
            if supplier_name:
                from ...models.supplier import Supplier
                sup = db.session.execute(
                    db.select(Supplier).filter(
                        db.func.lower(Supplier.name) == supplier_name.lower()
                    )
                ).scalar_one_or_none()
                if not sup:
                    # Auto-create supplier
                    sup = Supplier(name=supplier_name.strip())
                    db.session.add(sup)
                    db.session.flush()
                supplier_id = sup.id

            # Resolve/create category
            cat_val = category or None
            if category:
                existing_cat = db.session.execute(
                    db.select(Category).filter_by(name=category)
                ).scalar_one_or_none()
                if not existing_cat:
                    new_cat = Category(name=category)
                    db.session.add(new_cat)
                    db.session.flush()

            # Parse expiry
            expiry_date = None
            if expiry_raw:
                from datetime import date as _date, datetime, timezone
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
                    try:
                        expiry_date = _date.fromisoformat(expiry_raw) if fmt == "%Y-%m-%d" \
                            else _date.strptime(expiry_raw, fmt)
                        break
                    except ValueError:
                        continue

            # Check if product exists by SKU
            existing = db.session.execute(
                db.select(Product).filter_by(sku=sku)
            ).scalar_one_or_none()

            if existing:
                # Update existing product
                existing.name = name
                existing.category = cat_val
                existing.cost_price = cost
                existing.selling_price = sell
                existing.quantity = qty
                existing.unit = unit
                existing.supplier_id = supplier_id
                if expiry_date:
                    existing.expiry_date = expiry_date
                # Update new fields if provided
                barcode_val = r.get("barcode") or r.get("ean") or r.get("upc") or None
                if barcode_val:
                    existing.barcode = barcode_val
                max_disc = r.get("max_discount_pct") or r.get("max discount") or None
                if max_disc:
                    try:
                        existing.max_discount_pct = float(max_disc)
                    except ValueError:
                        pass
                tax_cat = r.get("tax_category") or r.get("tax category") or None
                if tax_cat:
                    existing.tax_category = tax_cat
                updated += 1
            else:
                # Create new product directly (no internal commit — we commit at end)
                try:
                    from sqlalchemy.exc import IntegrityError
                    barcode_val = r.get("barcode") or r.get("ean") or r.get("upc") or None
                    max_disc_val = None
                    try:
                        raw_md = r.get("max_discount_pct") or r.get("max discount") or ""
                        if raw_md:
                            max_disc_val = float(raw_md)
                    except ValueError:
                        pass
                    tax_cat_val = r.get("tax_category") or r.get("tax category") or "standard"
                    p_obj = Product(
                        name=name, category=cat_val, sku=sku,
                        cost_price=cost, selling_price=sell,
                        quantity=qty, unit=unit,
                        supplier_id=supplier_id, expiry_date=expiry_date,
                        barcode=barcode_val,
                        max_discount_pct=max_disc_val,
                        tax_category=tax_cat_val,
                    )
                    db.session.add(p_obj)
                    db.session.flush()  # catch duplicate SKU immediately
                    created += 1
                except Exception as e:
                    db.session.rollback()
                    errors.append(f"Row {row_num}: '{name}' skipped — {e}")
                    skipped += 1
                    continue

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f"Error saving data: {e}", "danger")
            return render_template("inventory/bulk_upload.html")

        # ── Auto-relink images after upload ──────────────────────────────────
        # When products are cleared and re-uploaded, old image files stay on
        # disk but lose their DB link. Scan the uploads folder and match any
        # image whose filename contains the product name (slug-style match).
        try:
            relinked = _relink_product_images()
            if relinked:
                flash(f"📸 Auto-linked {relinked} product image(s) from previous uploads.", "info")
        except Exception:
            pass

        flash(f"✅ Bulk upload complete — {created} created, {updated} updated, {skipped} skipped.", "success")
        for err in errors:
            flash(err, "warning")
        return redirect(url_for("inventory.list_products"))

    return render_template("inventory/bulk_upload.html")


@inventory_bp.route("/relink-images", methods=["POST"])
@admin_required
def relink_images():
    """Manually re-link product images from the uploads folder by name matching."""
    try:
        relinked = _relink_product_images()
        if relinked:
            flash(f"✅ Re-linked {relinked} product image(s) from uploads folder.", "success")
        else:
            flash("No unmatched images found — all products already have images or no matching files exist.", "info")
    except Exception as e:
        flash(f"Error re-linking images: {e}", "danger")
    return redirect(url_for("inventory.list_products"))




@inventory_bp.route("/autofill-all", methods=["POST"])
@admin_required
def autofill_all():
    """Bulk-autofill all products missing description or image."""
    limit = int(request.form.get("limit", 50))
    try:
        from ...services.product_autofill import autofill_all_empty
        results = autofill_all_empty(limit=limit)
        flash(
            f"Bulk autofill complete — {results['updated']} products updated, "
            f"{results['skipped']} skipped out of {results['total']} processed.",
            "success"
        )
    except Exception as exc:
        flash(f"Bulk autofill failed: {exc}", "danger")
    return redirect(url_for("inventory.list_products"))


# ── AI Content Generation API endpoints ──────────────────────────────────────

@inventory_bp.route("/api/ai/generate-description", methods=["POST"])
@admin_required
def api_ai_generate_description():
    """Generate AI product description. POST {name, category} → {description}"""
    from flask import jsonify
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    category = (data.get("category") or "").strip()
    if not name:
        return jsonify({"error": "Product name required"}), 400
    try:
        from ...services.gemini_client import gemini_generate, gemini_available
        if not gemini_available():
            return jsonify({"error": "GEMINI_API_KEY not configured"}), 503
        prompt = (
            f"Write a rich product description for '{name}' ({category or 'General'}) "
            f"sold at a premium dry fruits store in Nepal (GoldKernel). "
            f"2-3 paragraphs. Include health benefits with ✅ bullets and a 💡 How to Use tip. "
            f"150-220 words. Plain text only, no markdown headers."
        )
        result = gemini_generate(prompt, max_tokens=400)
        if not result:
            return jsonify({"error": "AI generation failed"}), 502
        return jsonify({"description": result.strip()})
    except Exception as exc:
        import logging
        logging.getLogger(__name__).error("AI description gen failed: %s", exc)
        return jsonify({"error": "Generation failed"}), 500


@inventory_bp.route("/api/ai/generate-tags", methods=["POST"])
@admin_required
def api_ai_generate_tags():
    """Generate AI product tags. POST {name, category, description} → {tags}"""
    from flask import jsonify
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    category = (data.get("category") or "").strip()
    description = (data.get("description") or "").strip()[:200]
    if not name:
        return jsonify({"error": "Product name required"}), 400
    try:
        from ...services.gemini_client import gemini_generate, gemini_available
        if not gemini_available():
            return jsonify({"error": "GEMINI_API_KEY not configured"}), 503
        prompt = (
            f"Generate 6-8 SEO-friendly product tags for: {name} (category: {category or 'General'}).\n"
            f"Context: {description}\n"
            f"Rules: lowercase, no spaces in tag, use hyphens for multi-word, relevant to Nepal market.\n"
            f"Reply with ONLY a JSON array: [\"tag1\", \"tag2\", \"tag3\"]"
        )
        result = gemini_generate(prompt, max_tokens=100, temperature=0.3)
        if not result:
            return jsonify({"error": "AI generation failed"}), 502
        import json, re
        m = re.search(r'\[.*?\]', result, re.DOTALL)
        if m:
            tags = json.loads(m.group())
            tags = [str(t).strip().lower() for t in tags if t]
            return jsonify({"tags": tags, "tags_string": ", ".join(tags)})
        return jsonify({"error": "Could not parse tags"}), 502
    except Exception as exc:
        import logging
        logging.getLogger(__name__).error("AI tags gen failed: %s", exc)
        return jsonify({"error": "Generation failed"}), 500


@inventory_bp.route("/api/ai/generate-meta", methods=["POST"])
@admin_required
def api_ai_generate_meta():
    """Generate AI SEO meta description + title. POST {name, category, description} → {meta_description, seo_title}"""
    from flask import jsonify
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    category = (data.get("category") or "").strip()
    description = (data.get("description") or "").strip()[:300]
    if not name:
        return jsonify({"error": "Product name required"}), 400
    try:
        from ...services.gemini_client import gemini_generate, gemini_available
        if not gemini_available():
            return jsonify({"error": "GEMINI_API_KEY not configured"}), 503
        prompt = (
            f"Write SEO content for a Nepal e-commerce product page.\n"
            f"Product: {name} | Category: {category or 'General'}\n"
            f"Description context: {description}\n\n"
            f"Reply ONLY with valid JSON (no markdown):\n"
            f'{{"meta_description": "under 155 chars, benefit-focused, includes Nepal/GoldKernel", '
            f'"seo_title": "under 60 chars, includes product name and buy/shop"}}'
        )
        result = gemini_generate(prompt, max_tokens=150, temperature=0.3)
        if not result:
            return jsonify({"error": "AI generation failed"}), 502
        import json, re
        # Strip markdown fences if present
        clean = re.sub(r"^```(?:json)?\s*", "", result.strip())
        clean = re.sub(r"\s*```$", "", clean)
        parsed = json.loads(clean)
        return jsonify({
            "meta_description": str(parsed.get("meta_description", ""))[:320],
            "seo_title": str(parsed.get("seo_title", ""))[:120],
        })
    except Exception as exc:
        import logging
        logging.getLogger(__name__).error("AI meta gen failed: %s", exc)
        return jsonify({"error": "Generation failed"}), 500


# ── Pexels Image Search API ───────────────────────────────────────────────────

@inventory_bp.route("/api/pexels/search", methods=["GET"])
@admin_required
def api_pexels_search():
    """
    Proxy Pexels search via server with browser-like headers to avoid Cloudflare blocks.
    GET ?q=almond&per_page=9
    """
    from flask import jsonify
    import urllib.request, urllib.parse, json as _json, urllib.error

    query = request.args.get("q", "").strip()
    per_page = min(int(request.args.get("per_page", 9)), 15)
    if not query:
        return jsonify({"error": "Query required"}), 400

    pexels_key = os.environ.get("PEXELS_API_KEY", "")
    if not pexels_key:
        return jsonify({"error": "PEXELS_API_KEY not configured"}), 503

    try:
        encoded = urllib.parse.quote(query)
        url = f"https://api.pexels.com/v1/search?query={encoded}&per_page={per_page}&orientation=square"
        req = urllib.request.Request(
            url,
            headers={
                "Authorization": pexels_key,
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36",
                "Accept": "application/json",
                "Accept-Language": "en-US,en;q=0.9",
                "Referer": "https://www.pexels.com/",
                "Origin": "https://www.pexels.com",
            }
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = _json.loads(resp.read())
        photos = [
            {
                "id": p["id"],
                "url": p["src"]["medium"],
                "full_url": p["src"]["large2x"],
                "thumb": p["src"]["small"],
                "photographer": p.get("photographer", ""),
                "alt": p.get("alt", query),
            }
            for p in data.get("photos", [])
        ]
        return jsonify({"photos": photos, "total": data.get("total_results", 0)})
    except urllib.error.HTTPError as e:
        err_body = e.read().decode()[:200]
        import logging
        logging.getLogger(__name__).warning("Pexels HTTP %s: %s", e.code, err_body)
        # Fallback: return curated hardcoded images for common product types
        return _pexels_curated_fallback(query, per_page)
    except Exception as exc:
        import logging
        logging.getLogger(__name__).error("Pexels search error: %s", exc)
        return _pexels_curated_fallback(query, per_page)


def _pexels_curated_fallback(query: str, limit: int = 9):
    """Return curated Pexels direct image URLs when API is blocked."""
    from flask import jsonify
    q = query.lower()
    CURATED = {
        "cashew":    [("4109080","Cashew nuts in bowl"),("4109081","Raw cashews"),("4109082","Cashew close up")],
        "almond":    [("6157052","Almonds pile"),("6157053","Almond nuts"),("1295572","Almond bowl")],
        "walnut":    [("3630197","Walnuts"),("3630198","Walnut halves"),("1295573","Walnuts pile")],
        "pistachio": [("5702716","Pistachios"),("5702717","Pistachio nuts"),("5702718","Green pistachio")],
        "raisin":    [("6157050","Raisins"),("6157051","Dried raisins"),("3650438","Dried grapes")],
        "date":      [("6157049","Medjool dates"),("6157048","Dates fruit"),("1640771","Date palm")],
        "fig":       [("4051347","Dried figs"),("4051348","Fig fruit"),("4051349","Figs bowl")],
        "apricot":   [("3644742","Dried apricots"),("3644743","Apricot"),("3644744","Orange apricot")],
        "peanut":    [("4110380","Peanuts"),("4110381","Groundnuts"),("4110382","Peanut pile")],
        "coconut":   [("1528051","Coconut"),("1528052","Coconut halved"),("1528053","Coconut pieces")],
    }
    matched = []
    for key, photos in CURATED.items():
        if key in q:
            matched = photos[:limit]
            break
    if not matched:
        matched = [("5632388","Dry fruits mix"),("4109080","Nuts assorted"),("6157052","Mixed nuts")]

    result = []
    for pid, alt in matched[:limit]:
        base = f"https://images.pexels.com/photos/{pid}/pexels-photo-{pid}.jpeg"
        result.append({
            "id": int(pid),
            "url": base + "?auto=compress&cs=tinysrgb&w=400",
            "full_url": base + "?auto=compress&cs=tinysrgb&w=1260",
            "thumb": base + "?auto=compress&cs=tinysrgb&w=200",
            "photographer": "Pexels",
            "alt": alt,
        })
    return jsonify({"photos": result, "total": len(result), "source": "curated"})


@inventory_bp.route("/api/pexels/import", methods=["POST"])
@admin_required
def api_pexels_import():
    """Import a Pexels image into Cloudinary (or local). POST {product_id, image_url, alt}"""
    from flask import jsonify
    import urllib.request
    data = request.get_json(silent=True) or {}
    product_id = data.get("product_id")
    image_url = (data.get("image_url") or "").strip()
    if not image_url or not image_url.startswith("https://"):
        return jsonify({"error": "Valid HTTPS image_url required"}), 400

    try:
        # Download image bytes
        req = urllib.request.Request(
            image_url,
            headers={"User-Agent": "SmartMart/1.0 (product image import)"}
        )
        with urllib.request.urlopen(req, timeout=15) as resp:
            img_bytes = resp.read()

        if len(img_bytes) < 5000:
            return jsonify({"error": "Image too small or download failed"}), 400

        # Try Cloudinary upload first
        from ...services.image_service import _cloudinary_available, _cloudinary, _local_upload_dir
        cl = _cloudinary()
        if cl and _cloudinary_available():
            import io
            result = cl.uploader.upload(
                io.BytesIO(img_bytes),
                folder="smartmart/products",
                transformation=[
                    {"width": 800, "height": 800, "crop": "limit", "quality": "auto:good"},
                ],
                resource_type="image",
            )
            identifier = "cld:" + result["public_id"]
            thumb_url = cl.CloudinaryImage(result["public_id"]).build_url(
                transformation=[{"width": 200, "height": 200, "crop": "fill", "quality": "auto"}]
            )
        else:
            # Local fallback
            import uuid, os
            filename = f"{uuid.uuid4().hex}.jpg"
            dest = os.path.join(_local_upload_dir(), filename)
            with open(dest, "wb") as f:
                f.write(img_bytes)
            identifier = filename
            thumb_url = f"/static/uploads/products/{filename}"

        # Update product if product_id provided
        if product_id:
            product = db.session.get(Product, int(product_id))
            if product:
                # Delete old image
                from ...services.image_service import delete_product_image
                if product.image_filename:
                    delete_product_image(product.image_filename)
                product.image_filename = identifier
                db.session.commit()

        return jsonify({"identifier": identifier, "thumb_url": thumb_url, "ok": True})

    except Exception as exc:
        import logging
        logging.getLogger(__name__).error("Pexels import failed: %s", exc)
        return jsonify({"error": "Import failed"}), 500

@inventory_bp.route("/export-csv")
@admin_required
def export_csv():
    """Export full product list as CSV."""
    import csv, io
    from flask import Response
    products = db.session.execute(db.select(Product).order_by(Product.name)).scalars().all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Name", "SKU", "Category", "Cost Price", "Selling Price",
                     "Quantity", "Unit", "Supplier", "Expiry Date"])
    for p in products:
        writer.writerow([
            p.name, p.sku, p.category or "",
            float(p.cost_price), float(p.selling_price),
            p.quantity, p.unit or "pcs",
            p.supplier.name if p.supplier else "",
            p.expiry_date.isoformat() if p.expiry_date else "",
        ])
    return Response(
        output.getvalue(), mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=inventory.csv"}
    )


@inventory_bp.route("/products/<int:product_id>/price", methods=["POST"])
@login_required
def record_supplier_price(product_id):
    _require_perm("can_edit_product")
    product = db.get_or_404(Product, product_id)
    try:
        from ...models.supplier_price_record import SupplierPriceRecord
        new_cost = Decimal(str(request.form.get("cost_price", 0) or 0 or 0))
        if new_cost <= 0:
            raise ValueError("Cost price must be greater than zero.")
        record = SupplierPriceRecord(
            product_id=product.id,
            supplier_name=request.form.get("supplier_name", "").strip() or None,
            cost_price=new_cost,
            quantity_kg=Decimal(str(request.form.get("quantity_kg", 0) or 0 or 0)) or None,
            invoice_ref=request.form.get("invoice_ref", "").strip() or None,
            recorded_by=current_user.id,
        )
        product.cost_price = new_cost
        db.session.add(record)
        db.session.commit()
        flash("Supplier price recorded and product cost updated.", "success")
    except Exception as exc:
        db.session.rollback()
        flash(f"Could not record supplier price: {exc}", "danger")
    return redirect(url_for("inventory.price_history", product_id=product_id))


@inventory_bp.route("/products/<int:product_id>/price-history")
@login_required
def price_history(product_id):
    _require_perm("can_view_stock_report")
    from ...models.supplier_price_record import SupplierPriceRecord

    product = db.get_or_404(Product, product_id)
    records = db.session.execute(
        db.select(SupplierPriceRecord)
        .where(SupplierPriceRecord.product_id == product.id)
        .order_by(SupplierPriceRecord.recorded_at)
    ).scalars().all()
    return render_template("inventory/price_history.html", product=product, records=records)


@inventory_bp.route("/price-alerts")
@login_required
def price_alerts():
    _require_perm("can_view_stock_report")
    from datetime import datetime, timedelta
    from ...models.supplier_price_record import SupplierPriceRecord

    cutoff = datetime.utcnow() - timedelta(days=90)
    products = db.session.execute(db.select(Product).order_by(Product.name)).scalars().all()
    alerts = []
    for product in products:
        latest = db.session.execute(
            db.select(SupplierPriceRecord)
            .where(SupplierPriceRecord.product_id == product.id)
            .order_by(SupplierPriceRecord.recorded_at.desc())
            .limit(1)
        ).scalar_one_or_none()
        old = db.session.execute(
            db.select(SupplierPriceRecord)
            .where(SupplierPriceRecord.product_id == product.id)
            .where(SupplierPriceRecord.recorded_at <= cutoff)
            .order_by(SupplierPriceRecord.recorded_at.desc())
            .limit(1)
        ).scalar_one_or_none()
        if latest and old and float(old.cost_price or 0) > 0:
            increase_pct = (float(latest.cost_price) - float(old.cost_price)) / float(old.cost_price) * 100
            if increase_pct >= 10:
                current_margin = 0.35
                suggested_price = float(latest.cost_price) / (1 - current_margin)
                alerts.append({
                    "product": product,
                    "increase_pct": increase_pct,
                    "old_cost": float(old.cost_price),
                    "new_cost": float(latest.cost_price),
                    "suggested_price": suggested_price,
                })
    return render_template("inventory/price_alerts.html", alerts=alerts)


@inventory_bp.route("/labels", methods=["GET", "POST"])
@login_required
def print_labels():
    """Generate printable barcode/price labels for products."""
    _require_perm("can_print_labels")
    products = db.session.execute(db.select(Product).order_by(Product.name)).scalars().all()
    selected_ids = []
    label_data = []

    if request.method == "POST":
        selected_ids = [int(x) for x in request.form.getlist("product_ids") if x.isdigit()]
        # Generate QR/barcode images as base64 for selected products
        for p in products:
            if p.id in selected_ids:
                barcode_b64 = _generate_barcode_b64(p.sku)
                label_data.append({
                    "id": p.id,
                    "name": p.name,
                    "sku": p.sku,
                    "price": float(p.selling_price),
                    "category": p.category or "",
                    "barcode_b64": barcode_b64,
                })

    from flask import jsonify as _jsonify
    shop = None
    try:
        shop = __import__('smart_mart.models.shop_settings', fromlist=['ShopSettings']).ShopSettings.get()
    except Exception:
        pass
    return render_template("inventory/labels.html",
                           products=products,
                           selected_ids=selected_ids,
                           label_data=label_data,
                           shop=shop)


@inventory_bp.route("/labels/barcode")
@login_required
def label_barcode():
    """Return a base64 Code128 barcode PNG for a given value (barcode or SKU)."""
    value = request.args.get("value", "").strip()
    if not value:
        from flask import jsonify as _jsonify
        return _jsonify({"barcode": ""})
    from flask import jsonify as _jsonify
    return _jsonify({"barcode": _generate_barcode_b64(value)})


@inventory_bp.route("/labels/pdf", methods=["POST"])
@login_required
def labels_pdf():
    """Generate a PDF where every page is exactly one label at the chosen size.
    The PDF can be opened in any PDF viewer and printed with:
      - Paper size = actual label size (no scaling)
      - Margins = none
    This bypasses browser CSS @page limitations entirely.
    """
    _require_perm("can_print_labels")
    from flask import Response as _Resp
    import io

    data        = request.get_json() or {}
    items       = data.get("items", [])
    w_mm        = float(data.get("w_mm", 75))
    h_mm        = float(data.get("h_mm", 50))
    show_bc     = data.get("show_barcode", True)
    show_shop   = data.get("show_shop", True)
    show_mrp    = data.get("show_mrp", True)
    show_sku    = data.get("show_sku", False)

    shop_name = ""
    try:
        from ...models.shop_settings import ShopSettings
        shop_name = ShopSettings.get().shop_name or ""
    except Exception:
        pass

    try:
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
        from reportlab.lib.utils import ImageReader

        W = w_mm * mm
        H = h_mm * mm
        pad = 2 * mm

        buf = io.BytesIO()
        c = canvas.Canvas(buf, pagesize=(W, H))

        for item in items:
            copies  = max(1, int(item.get("copies", 1)))
            name    = str(item.get("name", ""))
            sku     = str(item.get("sku", ""))
            bc_val  = str(item.get("barcode") or sku)
            price   = float(item.get("price", 0))
            price_str = f"NPR {int(price) if price == int(price) else f'{price:.2f}'}"

            # Generate barcode image
            bc_img = None
            if show_bc:
                try:
                    import barcode as _bc
                    from barcode.writer import ImageWriter
                    bb = io.BytesIO()
                    code = _bc.get("code128", bc_val, writer=ImageWriter())
                    code.write(bb, options={
                        "write_text": True,
                        "module_height": 8.0,
                        "module_width": 0.7,
                        "quiet_zone": 2.0,
                        "font_size": 6,
                        "text_distance": 1.0,
                    })
                    bb.seek(0)
                    bc_img = ImageReader(bb)
                except Exception:
                    bc_img = None

            for _ in range(copies):
                c.setPageSize((W, H))
                c.setFillColorRGB(0, 0, 0)

                # barcode at bottom
                bc_h = 12 * mm if bc_img else 0

                # ── Text layout (top down) ──
                # reportlab y=0 is bottom, so convert: y_from_top → H - y_from_top
                y_top = H - pad   # start from top minus padding

                def draw_text(text, font, size, y_ft):
                    """Draw text at y measured from top."""
                    c.setFont(font, size)
                    c.drawString(pad, y_ft - size * 0.352778, text)  # 1pt = 0.352778mm
                    return y_ft - (size * 0.352778 + 1.2 * mm)

                y = y_top
                if show_shop and shop_name:
                    y = draw_text(shop_name, "Helvetica-Bold", 5, y)

                # Product name — truncate to fit width
                max_w_mm = w_mm - 2 * 2  # subtract padding both sides
                font_size_name = 8
                c.setFont("Helvetica-Bold", font_size_name)
                char_w_mm = font_size_name * 0.352778 * 0.55
                max_chars = int(max_w_mm / char_w_mm)
                disp = name if len(name) <= max_chars else name[:max_chars - 1] + "…"
                y = draw_text(disp, "Helvetica-Bold", font_size_name, y)

                y = draw_text(price_str, "Helvetica-Bold", 12, y)

                if show_mrp:
                    y = draw_text("MRP incl. all taxes", "Helvetica", 4.5, y)

                if show_sku:
                    draw_text(sku, "Helvetica", 4.5, y)

                # Barcode spans full width at bottom
                if bc_img:
                    c.drawImage(bc_img, 0, 0, width=W, height=bc_h,
                                preserveAspectRatio=False)

                c.showPage()

        c.save()
        buf.seek(0)

        return _Resp(
            buf.read(),
            mimetype="application/pdf",
            headers={
                "Content-Disposition": "inline; filename=labels.pdf",
            }
        )

    except Exception as exc:
        import logging
        logging.getLogger(__name__).error("labels_pdf failed: %s", exc, exc_info=True)
        return {"error": str(exc)}, 500


@inventory_bp.route("/labels/print-direct", methods=["POST"])
@login_required
def labels_print_direct():
    """Print labels directly to the Windows sticker printer via GDI SetDIBitsToDevice."""
    _require_perm("can_print_labels")
    from flask import jsonify as _json
    import io, struct, ctypes

    data       = request.get_json() or {}
    items      = data.get("items", [])
    show_bc    = data.get("show_barcode", True)
    show_shop  = data.get("show_shop", True)
    show_mrp   = data.get("show_mrp", True)
    show_sku   = data.get("show_sku", False)
    printer_nm = data.get("printer", "sticker printer")

    shop_name = ""
    try:
        from ...models.shop_settings import ShopSettings
        shop_name = ShopSettings.get().shop_name or ""
    except Exception:
        pass

    try:
        import win32ui, win32con
        from PIL import Image, ImageDraw, ImageFont
        import barcode as _bc
        from barcode.writer import ImageWriter

        # XP-409B sticker: 1.97in x 0.98in = 50mm x 25mm
        # Draw in portrait orientation (200w x 400h) then rotate 90° clockwise
        # so the printer (which feeds portrait) gets the correct orientation.
        DPI   = 203
        PX_W  = int(25 / 25.4 * DPI)   # 200px — portrait width
        PX_H  = int(50 / 25.4 * DPI)   # 400px — portrait height

        # Open the printer DC — needed for print job lifecycle
        hdc = win32ui.CreateDC()
        hdc.CreatePrinterDC(printer_nm)

        _fp = r"C:\Windows\Fonts\arial.ttf"
        def _f(px):
            try:    return ImageFont.truetype(_fp, max(6, px))
            except: return ImageFont.load_default()

        # Font sizes for 200px wide portrait canvas
        f_shop  = _f(13)
        f_name  = _f(18)
        f_price = _f(26)
        f_small = _f(11)

        printed = 0
        hdc.StartDoc("Smart Mart Labels")

        for item in items:
            copies  = max(1, int(item.get("copies", 1)))
            name    = str(item.get("name", ""))
            sku     = str(item.get("sku", ""))
            bc_val  = str(item.get("barcode") or sku)
            price   = float(item.get("price", 0))
            price_s = f"NPR {int(price) if price == int(price) else f'{price:.2f}'}"

            for _ in range(copies):
                img  = Image.new("RGB", (PX_W, PX_H), (255, 255, 255))
                draw = ImageDraw.Draw(img)
                pad  = 5
                gap  = 2
                y    = pad

                def put(text, fnt):
                    nonlocal y
                    while len(text) > 1:
                        bb = draw.textbbox((pad, y), text, font=fnt)
                        if bb[2] <= PX_W - pad: break
                        text = text[:-1]
                    draw.text((pad, y), text, font=fnt, fill=(0, 0, 0))
                    bb = draw.textbbox((pad, y), text, font=fnt)
                    y += (bb[3] - bb[1]) + gap

                if show_shop and shop_name: put(shop_name, f_shop)
                put(name, f_name)
                put(price_s, f_price)
                if show_mrp: put("MRP incl. taxes", f_small)
                if show_sku: put(sku, f_small)

                # Barcode at bottom
                if show_bc:
                    try:
                        bb = io.BytesIO()
                        code = _bc.get("code128", bc_val, writer=ImageWriter())
                        code.write(bb, options={
                            "write_text": True, "module_height": 10.0,
                            "module_width": 0.4, "quiet_zone": 1.0,
                            "font_size": 5, "text_distance": 1.0, "dpi": 203,
                        })
                        bb.seek(0)
                        bc_im = Image.open(bb); bc_im.load()
                        bc_im = bc_im.convert("RGB")
                        bc_h  = int(PX_H * 0.38)   # ~9mm barcode on 25mm sticker
                        bc_im = bc_im.resize((PX_W, bc_h), Image.LANCZOS)
                        img.paste(bc_im, (0, PX_H - bc_h))
                    except Exception:
                        pass

                # The XP-409B feeds labels portrait (short side = feed direction).
                # Our image is landscape (400w x 200h) so rotate 90° clockwise
                # so it prints correctly when the printer feeds portrait.
                img_rotated = img.rotate(-90, expand=True)  # -90 = clockwise

                # Draw into printer DC — stretch rotated image to fill printer canvas
                hdc.StartPage()
                dc_w = hdc.GetDeviceCaps(win32con.HORZRES)
                dc_h = hdc.GetDeviceCaps(win32con.VERTRES)
                rw, rh = img_rotated.size
                raw    = img_rotated.convert("RGB").tobytes("raw", "BGR")
                bi_hdr = struct.pack("<lllHHLLllLL",
                    40, rw, -rh, 1, 24, 0, len(raw), 0, 0, 0, 0)
                ctypes.windll.gdi32.StretchDIBits(
                    hdc.GetSafeHdc(),
                    0, 0, dc_w, dc_h,
                    0, 0, rw, rh,
                    raw, bi_hdr,
                    0,
                    0x00CC0020
                )
                hdc.EndPage()
                printed += 1

        hdc.EndDoc()
        hdc.DeleteDC()
        return _json({"ok": True, "printed": printed})

    except ImportError as exc:
        return _json({"ok": False, "error": f"Missing: {exc}. Use Download PDF instead."}), 400
    except Exception as exc:
        import logging, traceback
        logging.getLogger(__name__).error("print_direct: %s\n%s", exc, traceback.format_exc())
        return _json({"ok": False, "error": str(exc)}), 500

def _generate_barcode_b64(value: str) -> str:
    """Generate a Code128 barcode as base64 PNG — scannable by any barcode scanner."""
    try:
        import barcode as _bc
        from barcode.writer import ImageWriter
        import io, base64
        buf = io.BytesIO()
        code = _bc.get(
            "code128", value,
            writer=ImageWriter()
        )
        code.write(buf, options={
            "write_text":    True,    # print the number below the bars
            "module_height": 8.0,     # bar height in mm
            "module_width":  0.8,     # bar width — narrower = more compact
            "quiet_zone":    2.0,     # whitespace either side
            "font_size":     6,       # text size below bars
            "text_distance": 1.0,
        })
        return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode()
    except Exception:
        # Fallback: QR code (always works, readable by phone but not handheld scanner)
        try:
            import qrcode, io, base64
            qr = qrcode.QRCode(version=1, box_size=3, border=1,
                               error_correction=qrcode.constants.ERROR_CORRECT_L)
            qr.add_data(value)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode()
        except Exception:
            return ""


@inventory_bp.route("/bulk-upload/sample")
@admin_required
def download_product_sample():
    """Download a sample CSV for bulk product upload."""
    from flask import Response
    sample = "Product Name,SKU,Category,Cost Price,Selling Price,Quantity,Unit,Supplier,Expiry Date\n"
    sample += "Basmati Rice,RICE-001,Grains & Pulses,80.00,120.00,50,kg,ABC Traders,2026-12-31\n"
    sample += "Mustard Oil 1L,OIL-001,Oils & Fats,150.00,200.00,30,pcs,XYZ Suppliers,\n"
    sample += "Colgate Toothpaste,TOOTH-001,Personal Care & Hygiene,75.00,110.00,20,pcs,,2027-06-30\n"
    sample += "Wai Wai Noodles,NOODLE-001,Snacks & Bakery,18.00,25.00,100,pcs,,\n"
    return Response(
        sample, mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=bulk_products_sample.csv"}
    )


# ── Product Variants ─────────────────────────────────────────────────────────

@inventory_bp.route("/<int:product_id>/variants")
@login_required
def product_variants(product_id):
    _require_perm("can_manage_variants")
    product = db.get_or_404(Product, product_id)
    from ...models.product_variant import ProductVariant
    variants = db.session.execute(
        db.select(ProductVariant).where(ProductVariant.product_id == product_id)
        .order_by(ProductVariant.variant_name)
    ).scalars().all()
    return render_template("inventory/variants.html", product=product, variants=variants)


@inventory_bp.route("/<int:product_id>/variants/create", methods=["GET", "POST"])
@login_required
def create_variant(product_id):
    _require_perm("can_manage_variants")
    product = db.get_or_404(Product, product_id)
    from ...models.product_variant import ProductVariant
    from sqlalchemy.exc import IntegrityError
    if request.method == "POST":
        variant_name = request.form.get("variant_name", "").strip()
        sku = request.form.get("sku", "").strip()
        cost_price = Decimal(str(request.form.get("cost_price", 0) or 0 or 0))
        selling_price = Decimal(str(request.form.get("selling_price", 0) or 0 or 0))
        quantity = int(request.form.get("quantity", 0) or 0)
        barcode = request.form.get("barcode", "").strip() or None
        if not variant_name or not sku:
            flash("Variant name and SKU are required.", "danger")
        else:
            try:
                v = ProductVariant(
                    product_id=product_id,
                    variant_name=variant_name,
                    sku=sku,
                    cost_price=cost_price,
                    selling_price=selling_price,
                    quantity=quantity,
                    barcode=barcode,
                    is_active=True,
                )
                db.session.add(v)
                db.session.commit()
                flash(f"Variant '{variant_name}' added.", "success")
                return redirect(url_for("inventory.product_variants", product_id=product_id))
            except IntegrityError:
                db.session.rollback()
                flash(f"SKU '{sku}' already exists.", "danger")
    return render_template("inventory/variant_form.html", product=product, variant=None, action="Add")


@inventory_bp.route("/<int:product_id>/variants/<int:variant_id>/edit", methods=["GET", "POST"])
@login_required
def edit_variant(product_id, variant_id):
    _require_perm("can_manage_variants")
    product = db.get_or_404(Product, product_id)
    from ...models.product_variant import ProductVariant
    from sqlalchemy.exc import IntegrityError
    variant = db.get_or_404(ProductVariant, variant_id)
    if request.method == "POST":
        variant.variant_name = request.form.get("variant_name", "").strip()
        variant.sku = request.form.get("sku", "").strip()
        variant.cost_price = Decimal(str(request.form.get("cost_price", 0) or 0 or 0))
        variant.selling_price = Decimal(str(request.form.get("selling_price", 0) or 0 or 0))
        variant.quantity = int(request.form.get("quantity", 0) or 0)
        variant.barcode = request.form.get("barcode", "").strip() or None
        variant.is_active = request.form.get("is_active") == "on"
        try:
            db.session.commit()
            flash("Variant updated.", "success")
            return redirect(url_for("inventory.product_variants", product_id=product_id))
        except IntegrityError:
            db.session.rollback()
            flash("SKU already exists.", "danger")
    return render_template("inventory/variant_form.html", product=product, variant=variant, action="Edit")


@inventory_bp.route("/<int:product_id>/variants/<int:variant_id>/delete", methods=["POST"])
@login_required
def delete_variant(product_id, variant_id):
    _require_perm("can_manage_variants")
    from ...models.product_variant import ProductVariant
    v = db.get_or_404(ProductVariant, variant_id)
    db.session.delete(v)
    db.session.commit()
    flash("Variant deleted.", "success")
    return redirect(url_for("inventory.product_variants", product_id=product_id))


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _form_to_data(form) -> dict:
    from datetime import date

    # Handle category: use selected or create new one
    category_val = form.get("category", "").strip() or None
    new_cat = form.get("new_category", "").strip()
    if new_cat:
        try:
            existing = db.session.execute(
                db.select(Category).filter_by(name=new_cat)
            ).scalar_one_or_none()
            if not existing:
                cat = Category(name=new_cat)
                db.session.add(cat)
                db.session.commit()
        except Exception:
            db.session.rollback()
        category_val = new_cat
    elif category_val == "__new__":
        category_val = None

    data: dict = {
        "name": form.get("name", "").strip(),
        "category": category_val,
        "sku": form.get("sku", "").strip(),
        "cost_price": form.get("cost_price", "0") or "0",
        "selling_price": form.get("selling_price", "0") or "0",
        "quantity": int(form.get("quantity", 0) or 0),
        "supplier_id": int(form.get("supplier_id")) if form.get("supplier_id") else None,
        "purchase_date": None,
        "expiry_date": None,
        "unit": form.get("unit", "pcs").strip() or "pcs",
        "reorder_point": int(form.get("reorder_point", 10) or 10),
        "low_stock_threshold": int(form.get("low_stock_threshold", 500) or 500),
        "barcode": form.get("barcode", "").strip() or None,
        "is_active": form.get("is_active") == "on",
        "tax_category": form.get("tax_category", "standard").strip() or "standard",
        "max_discount_pct": None,
        # ── Store content fields (manual edit — see inventory/form.html) ─────
        "description":       form.get("description", "").strip() or None,
        "benefits":          form.get("benefits", "").strip() or None,
        "origin":            form.get("origin", "").strip() or None,
        "storage_tips":      form.get("storage_tips", "").strip() or None,
        # ── SEO & Discovery fields ────────────────────────────────────────────
        "tags":              form.get("tags", "").strip() or None,
        "meta_description":  form.get("meta_description", "").strip()[:320] or None,
        "seo_title":         form.get("seo_title", "").strip()[:120] or None,
    }
    # Purchase date
    purchase_date_raw = form.get("purchase_date", "").strip()
    if purchase_date_raw:
        try:
            data["purchase_date"] = date.fromisoformat(purchase_date_raw)
        except ValueError:
            pass
    max_disc_raw = form.get("max_discount_pct", "").strip()
    if max_disc_raw:
        try:
            data["max_discount_pct"] = float(max_disc_raw)
        except ValueError:
            pass
    expiry_raw = form.get("expiry_date", "").strip()
    if expiry_raw:
        try:
            data["expiry_date"] = date.fromisoformat(expiry_raw)
        except ValueError:
            pass
    return data
